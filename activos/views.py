import os
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Activo, DEPARTAMENTOS, TransferenciaActivo
from django.db.models import Q, Count
from .forms import ActivoForm
from django import forms
from django.urls import reverse
from django.http import HttpResponse, HttpResponseForbidden
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.drawing.image import Image
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.conf import settings
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.http import HttpResponse
import os
from datetime import datetime
from .models import Activo
from activos import models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import json
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.views import PasswordResetView, PasswordResetConfirmView
from django.contrib.auth.forms import PasswordResetForm, SetPasswordForm



# 📍login 
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, 'activos/login.html')

@login_required
# 📍 Registrar activo
def registrar_activo(request, departamento=None):
    if request.method == 'POST':
        form = ActivoForm(request.POST)
        if form.is_valid():
            # Guarda el formulario pero NO establezcas departamento aquí
            activo = form.save(commit=False)
            
            # Asegúrate de asignar el departamento
            if departamento:
                activo.departamento = departamento
                activo.save()
            
            return redirect('detalle_departamento', departamento=activo.departamento)
    else:

        form = ActivoForm(initial={'activo': False})

    return render(request, 'activos/registrar_activo.html', {
        'form': form, 
        'departamento': departamento
    })

@login_required
# 📍 Editar activo
def editar_activo(request, pk):
    activo = get_object_or_404(Activo, pk=pk)
    departamento = activo.departamento

    if request.method == 'POST':
        form = ActivoForm(request.POST, instance=activo)
        if form.is_valid():
            form.save()
            return redirect(reverse('detalle_departamento', args=[departamento]))
    else:
        form = ActivoForm(instance=activo)

    return render(request, 'activos/editar_activo.html', {'form': form, 'departamento': departamento})


@login_required
#📍 dash
def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # Lista de departamentos fijos - CORREGIDA (faltaba coma)
    departamentos = [
        "Fiscalización",
        "Recaudación", 
        "Inmuebles Urbanos",
        "Gerencia de Licores",
        "Gerencia General",
        "Jurídica",
        "Informatica",  # CORREGIDO: agregada coma
        "Administración y Finanzas"
    ]

    # Si el usuario es admin → ve todo
    if request.user.is_superuser:
        activos = Activo.objects.all()
    else:
        # Usuario normal → solo su departamento
        activos = Activo.objects.filter(departamento=request.user.departamento)

    # Conteo total - CORREGIDO: usar minúsculas
    total_activos = activos.count()
    activos_operativos = activos.filter(condicion="operativo").count()  # minúscula
    activos_danados = activos.filter(condicion="dañado").count()        # minúscula
    

    # Agrupar por departamento (solo para admin)
    if request.user.is_superuser:
        activos_db = activos.values('departamento').annotate(total=Count('id'))
        activos_dict = {item['departamento']: item['total'] for item in activos_db}

        activos_por_departamento = [
            {"departamento": depto, "total": activos_dict.get(depto, 0)}
            for depto in departamentos
        ]
    else:
        activos_por_departamento = [
            {"departamento": request.user.departamento, "total": total_activos}
        ]

    context = {
        "total_activos": total_activos,
        "activos_operativos": activos_operativos,
        "activos_danados": activos_danados,
             "activos_por_departamento": activos_por_departamento,
        "labels": [d["departamento"] for d in activos_por_departamento],
        "data": [d["total"] for d in activos_por_departamento],
    }
    return render(request, "activos/dashboard.html", context)

@login_required
#📍 dash departamento
def detalle_departamento(request, departamento):
    # Filtrar los activos solo del departamento seleccionado (por nombre)
    activos = Activo.objects.filter(departamento=departamento)

    # Contadores
    total_activos = activos.count()
    activos_operativos = activos.filter(condicion__icontains='Operativo').count()
    activos_danados = activos.filter(condicion__icontains='Dañado').count()
    

    # Contexto
    context = {
        'departamento': departamento,  # ahora es texto, no objeto
        'activos': activos,
        'total_activos': total_activos,
        'activos_operativos': activos_operativos,
        'activos_danados': activos_danados,
    }

    return render(request, 'activos/detalle_departamento.html', context)

@login_required
# 📍 Cerrar seccion
def logout_view(request):

    logout(request)
    return redirect('login')

@login_required
# 📍 exportar excel todos 
def exportar_activos(request):
   # Crear el libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario General de Bienes"

    

    # -------- ENCABEZADO --------
    ws["B1"] = "Organismo:"
    ws["C1"] = "0001 - ALCALDÍA DE IRIBARREN"
    ws["F1"] = f"FECHA: {datetime.now().strftime('%d/%m/%Y')}"

    ws["B2"] = "Servicio:"
    ws["C2"] = "SERVICIO MUNICIPAL DE ADMINISTRACIÓN TRIBUTARIA"
    ws["B3"] = "Unidad Administrativa:"
    ws["C3"] = "SERVICIO MUNICIPAL DE ADMINISTRACIÓN TRIBUTARIA"
    ws["A4"] = "Distrito:"
    ws["B4"] = "LARA"
    ws["E4"] = "INVENTARIO GENERAL DE BIENES"
    ws["E4"].font = Font(bold=True)
    ws["A5"] = "Dirección o Lugar:"
    ws["B5"] = "CALLE 26 ENTRE CARRERAS 15 Y 16"

    ws.append([])  # espacio antes de la tabla

    # -------- ENCABEZADOS DE TABLA --------
    encabezados = ['Responsable', 'Código', 'Serial', 'Descripción', 'Estado', 'Departamento']
    ws.append(encabezados)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[7]:
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

           # -------- DATOS AGRUPADOS POR RESPONSABLE --------
    activos = Activo.objects.all().order_by('responsable')

    # Agrupar por responsable
    datos = {}
    for a in activos:
        datos.setdefault(a.responsable, []).append(a)

    fila = 8  # primera fila de datos
    for responsable, bienes in datos.items():
        primera_fila = True  # bandera para saber si es la primera del grupo

        for bien in bienes:
            ws.append([
                responsable if primera_fila else '',  # solo muestra el nombre una vez
                bien.codigo,
                bien.serial,
                bien.descripcion,
                bien.get_condicion_display(),
                bien.departamento if bien.departamento else '',
            ])
            primera_fila = False  # las siguientes filas quedan vacías en responsable
            fila += 1


    # -------- BORDES Y FORMATOS --------
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # -------- ANCHO DE COLUMNAS --------
    ancho_uniforme = 22
    for col in range(1, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho_uniforme

    # -------- RESPUESTA --------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Inventario_General.xlsx'
    wb.save(response)
    return response

@login_required
#📍 exportar excel por departamento 
def exportar_activos_departamento(request, departamento):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Inventario {departamento}"


    # -------- ENCABEZADO --------
    ws["B1"] = "Organismo:"
    ws["C1"] = "0001 - ALCALDÍA DE IRIBARREN"
    ws["F1"] = f"FECHA: {datetime.now().strftime('%d/%m/%Y')}"
    ws["B2"] = "Servicio:"
    ws["C2"] = "SERVICIO MUNICIPAL DE ADMINISTRACIÓN TRIBUTARIA"
    ws["B3"] = "Unidad Administrativa:"
    ws["C3"] = "SERVICIO MUNICIPAL DE ADMINISTRACIÓN TRIBUTARIA"
    ws["A4"] = "Distrito:"
    ws["B4"] = "LARA"
    ws["E4"] = "INVENTARIO DE BIENES POR DEPARTAMENTO"
    ws["E4"].font = Font(bold=True)
    ws["A5"] = "Departamento:"
    ws["B5"] = departamento
    ws["A6"] = "Dirección o Lugar:"
    ws["B6"] = "CALLE 26 ENTRE CARRERAS 15 Y 16"

    ws.append([])

    # -------- ENCABEZADOS DE TABLA --------
    encabezados = ['Responsable', 'Código', 'Serial', 'Descripción', 'Estado']
    ws.append(encabezados)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[8]:
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # -------- DATOS AGRUPADOS POR RESPONSABLE --------
    activos = Activo.objects.filter(departamento=departamento).order_by('responsable')
    datos = {}
    for a in activos:
        datos.setdefault(a.responsable, []).append(a)

    for responsable, bienes in datos.items():
        primera_fila = True
        for bien in bienes:
            ws.append([
                responsable if primera_fila else '',
                bien.codigo,
                bien.serial,
                bien.descripcion,
                bien.get_condicion_display()
            ])
            primera_fila = False

    # -------- FORMATO --------
    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in range(1, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # -------- RESPUESTA --------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"Inventario_{departamento.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
    wb.save(response)
    return response

@login_required
# 📍 Lista de activos (operativos o dañados)
def lista_activos(request):
    departamento = request.GET.get("depto")
    estado = request.GET.get("estado")

    # Iniciar con todos los activos (activos=True)
    activos = Activo.objects.filter(activo=True)
    total_activos = activos.count()
    activos_operativos = activos.filter(condicion="operativo").count()  # minúscula
    activos_danados = activos.filter(condicion="dañado").count()        # minúscula
  
    if departamento:
        activos = activos.filter(departamento=departamento)

    # CORRECCIÓN: Usar los valores exactos del modelo en minúsculas
    if estado:
        if estado == "dañado":
            activos = activos.filter(condicion="dañado")  # minúscula
        elif estado == "operativo":
            activos = activos.filter(condicion="operativo")  # minúscula

    

    context = {
        'activos': activos,
        'departamento': departamento,
        'estado': estado,
    }
    return render(request, 'activos/lista_activos.html', context)

@login_required
def eliminar_activo(request, pk):
    activo = get_object_or_404(Activo, pk=pk)
    departamento = activo.departamento

    if request.method == "POST":
        activo.condicion = 'dañado'      
        activo.activo = True       
        activo.save()
        
        
        # Redirección según usuario
        if request.user.is_superuser:
            return redirect('lista_activos')
        else:
            return redirect(reverse('activos_departamento', args=[departamento]))

    return render(request, 'activos/eliminar_activo.html', {
        'activo': activo,
        'departamento': departamento
    })

@login_required
#📍 Lista de activos (operativos o dañados) DEPARTAMENTOS
def activos_departamento(request, departamento):
    activos = Activo.objects.filter(departamento=departamento, activo=True)

    estado = request.GET.get('estado')
    buscar = request.GET.get('buscar')

    if buscar:
        activos = activos.filter(
        Q(serial__icontains=buscar) |
        Q(descripcion__icontains=buscar)
    )   
    # Filtro por estado
    if estado:
        if estado == "dañado":
            activos = activos.filter(condicion="dañado")  # minúscula
        elif estado == "operativo":
            activos = activos.filter(condicion="operativo")  # minúscula


    context = {
        "activos": activos,
        "departamento": departamento,
        "estado": estado
    }

    return render(request, "activos/activos_departamento.html", context)

@login_required
#📍 exportar pdf todos
def exportar_pdf_activos(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Inventario_General.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, 
                            topMargin=40, bottomMargin=40, 
                            leftMargin=40, rightMargin=40)
    elements = []
    estilos = getSampleStyleSheet()
    
    estilo_centrado = ParagraphStyle(
        name='Centrado',
        parent=estilos['Normal'],
        alignment=TA_CENTER,
        fontSize=10,
        fontName='Times-Roman'
    )

    # Cargar logos
    logo1_path = os.path.join(settings.BASE_DIR, "activos/static/activos/img/logo1.jpg")
    logo2_path = os.path.join(settings.BASE_DIR, "activos/static/activos/img/slider1.jpg")

    logo1 = Image(logo1_path, width=70, height=70) if os.path.exists(logo1_path) else Paragraph("LOGO 1", estilo_centrado)
    logo2 = Image(logo2_path, width=70, height=70) if os.path.exists(logo2_path) else Paragraph("LOGO 2", estilo_centrado)

    texto_encabezado = Paragraph("""
        <b>República Bolivariana de Venezuela</b><br/>
        <b>Estado Lara</b><br/>
        <b>Alcaldía Bolivariana del Municipio Iribarren</b><br/>
        <b>Servicio Municipal de Administración Tributaria</b><br/>
        <b>SEMAT</b>
    """, estilo_centrado)

    encabezado = Table([[logo1, texto_encabezado, logo2]], colWidths=[80, 400, 80])
    encabezado.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(encabezado)
    elements.append(Spacer(1, 20))

    # Estilos personalizados
    estilos.add(ParagraphStyle(name='TituloPrincipal', parent=estilos['Heading1'],
                               fontSize=16, textColor=colors.black, alignment=TA_CENTER,
                               spaceAfter=20, spaceBefore=10, fontName='Times-Bold'))

    estilos.add(ParagraphStyle(name='CeldaNormal', parent=estilos['Normal'],
                               fontSize=8, textColor=colors.black, alignment=TA_CENTER,
                               fontName='Times-Roman'))

    estilos.add(ParagraphStyle(name='CeldaEncabezado', parent=estilos['Normal'],
                               fontSize=9, textColor=colors.black, alignment=TA_CENTER,
                               fontName='Times-Bold'))

    # Título principal
    elements.append(Paragraph("INVENTARIO GENERAL DE BIENES", estilos['TituloPrincipal']))
    elements.append(Spacer(1, 15))

    # Tabla de datos
    encabezados = ['Nº', 'Responsable', 'Codigo', 'Serial','Descripción','Categoria','Subcategoria', 'Estado', 'Departamento']
    data = [encabezados]

    activos = Activo.objects.all().order_by("codigo")

    for idx, activo in enumerate(activos, 1):
        # Estado
        if activo.condicion == "operativo":
            estado_text = "Operativo"
        elif activo.condicion == "dañado":
            estado_text = "Dañado"
        else:
            estado_text = activo.get_condicion_display()

        # Categoría y subcategoría
        categoria_text = activo.categoria_principal.replace('_', ' ').title() if activo.categoria_principal else ""
        subcategoria_text = activo.subcategoria.replace('_', ' ').title() if activo.subcategoria else ""

        fila = [
            Paragraph(str(idx), estilos['CeldaNormal']),
            Paragraph(activo.responsable or "", estilos['CeldaNormal']),
            Paragraph(activo.codigo or "", estilos['CeldaNormal']),
            Paragraph(activo.serial or "", estilos['CeldaNormal']),
            Paragraph(activo.descripcion or "", estilos['CeldaNormal']),
            Paragraph(categoria_text, estilos['CeldaNormal']),
            Paragraph(subcategoria_text, estilos['CeldaNormal']),
            Paragraph(estado_text, estilos['CeldaNormal']),
            Paragraph(activo.departamento or "", estilos['CeldaNormal']),
        ]
        data.append(fila)

    # Si no hay activos, agregar filas vacías
    if not activos:
        for _ in range(3):
            fila_vacia = [Paragraph("", estilos['CeldaNormal']) for _ in range(9)]
            data.append(fila_vacia)

    tabla = Table(data, repeatRows=1, colWidths=[20, 60, 50, 60, 60, 70, 70, 40, 60])
    tabla.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ]))
    elements.append(tabla)
    elements.append(Spacer(1, 20))

    # Pie de página
    def pie_pagina(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 10)
        canvas.setFillColor(colors.black)
        canvas.drawString(40, 25, "SEMAT - Sistema de Gestión de Activos")
        canvas.drawRightString(doc.pagesize[0] - 40, 30, str(doc.page))
        canvas.restoreState()

    doc.build(elements, onFirstPage=pie_pagina, onLaterPages=pie_pagina)

    return response

@login_required
def exportar_pdf_activos_departamento(request, departamento):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Inventario_{departamento}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, 
                            topMargin=40, bottomMargin=40, 
                            leftMargin=40, rightMargin=40)
    elements = []
    estilos = getSampleStyleSheet()
    
    estilo_centrado = ParagraphStyle(
        name='Centrado',
        parent=estilos['Normal'],
        alignment=TA_CENTER,
        fontSize=10,
        fontName='Times-Roman'
    )

    # Cargar logos
    logo1_path = os.path.join(settings.BASE_DIR, "activos/static/activos/img/logo1.jpg")
    logo2_path = os.path.join(settings.BASE_DIR, "activos/static/activos/img/slider1.jpg")

    logo1 = Image(logo1_path, width=70, height=70) if os.path.exists(logo1_path) else Paragraph("LOGO 1", estilo_centrado)
    logo2 = Image(logo2_path, width=70, height=70) if os.path.exists(logo2_path) else Paragraph("LOGO 2", estilo_centrado)

    texto_encabezado = Paragraph("""
        <b>República Bolivariana de Venezuela</b><br/>
        <b>Estado Lara</b><br/>
        <b>Alcaldía Bolivariana del Municipio Iribarren</b><br/>
        <b>Servicio Municipal de Administración Tributaria</b><br/>
        <b>SEMAT</b>
    """, estilo_centrado)

    encabezado = Table([[logo1, texto_encabezado, logo2]], colWidths=[80, 400, 80])
    encabezado.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(encabezado)
    elements.append(Spacer(1, 20))

    # Estilos personalizados
    estilos.add(ParagraphStyle(name='TituloPrincipal', parent=estilos['Heading1'],
                               fontSize=16, textColor=colors.black, alignment=TA_CENTER,
                               spaceAfter=20, spaceBefore=10, fontName='Times-Bold'))

    estilos.add(ParagraphStyle(name='Subtitulo', parent=estilos['Normal'],
                               fontSize=12, textColor=colors.black, alignment=TA_CENTER,
                               spaceAfter=15, fontName='Times-Bold'))

    estilos.add(ParagraphStyle(name='InfoDepartamento', parent=estilos['Normal'],
                               fontSize=10, textColor=colors.black, alignment=TA_LEFT,
                               spaceAfter=10, fontName='Times-Roman'))

    estilos.add(ParagraphStyle(name='CeldaNormal', parent=estilos['Normal'],
                               fontSize=8, textColor=colors.black, alignment=TA_CENTER,
                               fontName='Times-Roman'))

    estilos.add(ParagraphStyle(name='CeldaEncabezado', parent=estilos['Normal'],
                               fontSize=9, textColor=colors.black, alignment=TA_CENTER,
                               fontName='Times-Bold'))

    # Título principal
    elements.append(Paragraph("INVENTARIO DE BIENES POR DEPARTAMENTO", estilos['TituloPrincipal']))
    elements.append(Spacer(1, 10))

    # Información del departamento
    info_table_data = [
        [Paragraph("<b>Departamento:</b>", estilos['InfoDepartamento']), 
         Paragraph(departamento, estilos['InfoDepartamento'])],
        [Paragraph("<b>Organismo:</b>", estilos['InfoDepartamento']), 
         Paragraph("0001 - ALCALDÍA DE IRIBARREN", estilos['InfoDepartamento'])],
        [Paragraph("<b>Unidad Administrativa:</b>", estilos['InfoDepartamento']), 
         Paragraph("SERVICIO MUNICIPAL DE ADMINISTRACIÓN TRIBUTARIA", estilos['InfoDepartamento'])],
        [Paragraph("<b>Dirección:</b>", estilos['InfoDepartamento']), 
         Paragraph("CALLE 26 ENTRE CARRERAS 15 Y 16", estilos['InfoDepartamento'])],
        [Paragraph("<b>Fecha:</b>", estilos['InfoDepartamento']), 
         Paragraph(datetime.now().strftime('%d/%m/%Y'), estilos['InfoDepartamento'])],
    ]

    info_table = Table(info_table_data, colWidths=[120, 400])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))

    # Tabla de datos
    encabezados = ['Nº', 'Responsable', 'Código', 'Serial', 'Descripción', 'Estado']
    data = [encabezados]

    # Obtener activos del departamento
    activos = Activo.objects.filter(departamento=departamento).order_by('responsable')
    
    # Agrupar por responsable (similar a tu función Excel)
    datos_agrupados = {}
    for a in activos:
        datos_agrupados.setdefault(a.responsable, []).append(a)

    idx = 1
    for responsable, bienes in datos_agrupados.items():
        primera_fila = True  # Para mostrar responsable solo en primera fila
        
        for bien in bienes:
            # Estado
            if bien.condicion == "operativo":
                estado_text = "Operativo"
            elif bien.condicion == "dañado":
                estado_text = "Dañado"
            else:
                estado_text = bien.get_condicion_display()

            fila = [
                Paragraph(str(idx), estilos['CeldaNormal']),
                Paragraph(responsable if primera_fila else "", estilos['CeldaNormal']),
                Paragraph(bien.codigo or "", estilos['CeldaNormal']),
                Paragraph(bien.serial or "", estilos['CeldaNormal']),
                Paragraph(bien.descripcion or "", estilos['CeldaNormal']),
                Paragraph(estado_text, estilos['CeldaNormal']),
            ]
            data.append(fila)
            idx += 1
            primera_fila = False  # Las siguientes filas quedan vacías en responsable

    # Si no hay activos, agregar filas vacías
    if not activos:
        for _ in range(3):
            fila_vacia = [Paragraph("", estilos['CeldaNormal']) for _ in range(6)]
            data.append(fila_vacia)

    tabla = Table(data, repeatRows=1, colWidths=[20, 80, 50, 60, 100, 50])
    tabla.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ]))
    elements.append(tabla)
    elements.append(Spacer(1, 20))

    # Resumen
    total_activos = activos.count()
    elementos_resumen = [
        Paragraph(f"<b>RESUMEN DEL DEPARTAMENTO: {departamento}</b>", estilos['Subtitulo']),
        Spacer(1, 10),
        Paragraph(f"<b>Total de bienes registrados:</b> {total_activos}", estilos['InfoDepartamento']),
    ]
    for elemento in elementos_resumen:
        elements.append(elemento)

    elements.append(Spacer(1, 30))

    # Pie de página
    def pie_pagina(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(colors.black)
        canvas.drawString(40, 25, f"SEMAT - Inventario de {departamento}")
        canvas.drawRightString(doc.pagesize[0] - 40, 30, f"Página {doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=pie_pagina, onLaterPages=pie_pagina)

    return response


def get_user_department(user):
    """
    Obtiene el departamento del usuario actual.
    Retorna el departamento o None si es admin.
    """
    # Si es superusuario, ve todo
    if user.is_superuser:
        return None
    
    # Obtener departamento directamente del campo
    if hasattr(user, 'departamento') and user.departamento:
        return user.departamento
    
    # Si no tiene departamento, verificar en grupos
    grupos_departamento = {
        'Informatica': 'Informatica',
        'Fiscalizacion': 'Fiscalización',
        'Recaudacion': 'Recaudación',
        'Juridica': 'Jurídica',
        'Administracion': 'Administración y Finanzas',
        'Gerencia': 'Gerencia General',
        'Inmuebles': 'Inmuebles Urbanos',
        'Licores': 'Gerencia de Licores',
    }
    
    for group_name, depto_name in grupos_departamento.items():
        if user.groups.filter(name=group_name).exists():
            # Actualizar el campo departamento para futuras consultas
            user.departamento = depto_name
            user.save()
            return depto_name
    
    # Si llegó aquí, no tiene acceso
    return "NO_ASIGNADO"

def filter_by_user_permission(queryset, user):
    """
    Filtra un queryset según los permisos del usuario.
    """
    if user.is_superuser:
        return queryset
    
    user_depto = get_user_department(user)
    
    if user_depto == "NO_ASIGNADO":
        return queryset.none()
    
    if user_depto:
        return queryset.filter(departamento=user_depto)
    
    return queryset.none()


@login_required
# 📍 Cambiar contraseña (cuando ya estás logueado)
def simple_password_change(request):
   
    if not request.user.is_authenticated:
        messages.error(request, "Debes iniciar sesión primero")
        return redirect('login')
    
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validaciones
        errors = []
        
        # 1. Verificar que la contraseña actual sea correcta
        if not request.user.check_password(current_password):
            errors.append("La contraseña actual es incorrecta")
        
        # 2. Verificar que las nuevas contraseñas coincidan
        if new_password != confirm_password:
            errors.append("Las nuevas contraseñas no coinciden")
        
        # 3. Validar longitud mínima
        if len(new_password) < 6:
            errors.append("La nueva contraseña debe tener al menos 6 caracteres")
        
        # 4. Validar que no sea igual a la actual
        if current_password == new_password:
            errors.append("La nueva contraseña no puede ser igual a la actual")
        
        # Si hay errores, mostrarlos
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Cambiar la contraseña
                request.user.set_password(new_password)
                request.user.save()
                
                # Mantener la sesión activa
                update_session_auth_hash(request, request.user)
                
                
                # Redirigir según el tipo de usuario
                if request.user.is_superuser:
                    return redirect('dashboard')
                else:
                    user_depto = get_user_department(request.user)
                    if user_depto and user_depto != "NO_ASIGNADO":
                        return redirect('detalle_departamento', departamento=user_depto)
                    else:
                        return redirect('dashboard')
                        
            except Exception as e:
                messages.error(request, f"Error al cambiar la contraseña: {str(e)}")
    
    return render(request, 'activos/cambiar_password.html')

# 📍 Solicitar recuperación de contraseña (olvidé mi contraseña)
def password_reset_request(request):
   
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        username = request.POST.get('username')
        
        if email or username:
            try:
                # Buscar usuario por email o username
                if email:
                    user = User.objects.get(email=email)
                else:
                    user = User.objects.get(username=username)
                
                # Generar token
                token = default_token_generator.make_token(user)
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                
                # Crear enlace de recuperación
                reset_url = request.build_absolute_uri(
                    f'/reset-password-confirm/{uid}/{token}/'
                )
                
                # Enviar email (simulado en desarrollo)
                subject = 'Recuperación de contraseña - SEMAT'
                message = f"""
                Hola {user.username},
                
                Has solicitado recuperar tu contraseña del Sistema SEMAT.
                
                Para restablecer tu contraseña, haz clic en el siguiente enlace:
                {reset_url}
                
                Este enlace expirará en 24 horas.
                
                Si no solicitaste este cambio, ignora este mensaje.
                
                Atentamente,
                Equipo SEMAT
                """
                
                # En producción, descomenta esto:
                # send_mail(
                #     subject,
                #     message,
                #     settings.DEFAULT_FROM_EMAIL,
                #     [user.email],
                #     fail_silently=False,
                # )
                
                # En desarrollo, mostramos el enlace
                print(f"\n🔐 ENLACE DE RECUPERACIÓN (para desarrollo):")
                print(f"   Usuario: {user.username}")
                print(f"   Email: {user.email}")
                print(f"   Enlace: {reset_url}")
                print("-" * 80)
                
                messages.info(request, 
                    f"Para desarrollo: El enlace es: {reset_url}")
                
                return redirect('login')
                
            except User.DoesNotExist:
                messages.error(request, 
                    "No se encontró un usuario con ese email o nombre de usuario.")
            except Exception as e:
                messages.error(request, f"Error al procesar la solicitud: {str(e)}")
        else:
            messages.error(request, "Debes ingresar un email o nombre de usuario")
    
    return render(request, 'activos/password_reset_request.html')

# 📍 Confirmar recuperación de contraseña
def password_reset_confirm(request, uidb64=None, token=None):
    """
    Vista para confirmar y establecer nueva contraseña.
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    try:
        # Decodificar el UID
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    # Verificar token
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            # Validaciones
            if new_password != confirm_password:
                messages.error(request, "Las contraseñas no coinciden")
            elif len(new_password) < 6:
                messages.error(request, "La contraseña debe tener al menos 6 caracteres")
            else:
                try:
                    # Establecer nueva contraseña
                    user.set_password(new_password)
                    user.save()
                    
                    messages.success(request, 
                        "¡Contraseña restablecida exitosamente! Ahora puedes iniciar sesión.")
                    return redirect('login')
                    
                except Exception as e:
                    messages.error(request, f"Error al cambiar la contraseña: {str(e)}")
        
        return render(request, 'activos/password_reset_confirm.html', {
            'validlink': True,
            'uidb64': uidb64,
            'token': token
        })
    else:
        messages.error(request, 
            "El enlace de recuperación no es válido o ha expirado.")
        return redirect('password_reset_request')
    
@login_required
def restaurar_activo(request, activo_id):
    print(f"DEBUG: Entrando a restaurar_activo con ID: {activo_id}")
    print(f"DEBUG: Método: {request.method}")
    
    if request.method == 'POST':
        activo = get_object_or_404(Activo, id=activo_id)
        
        if activo.condicion.lower() == 'dañado':  
            activo.condicion = 'operativo'  
            activo.save()
        else:
            print(f"DEBUG: Condición actual es '{activo.condicion}', no 'dañado'")
            messages.warning(request, f'No se puede restaurar. El activo está {activo.condicion}.')
    
    return redirect('lista_activos')


@login_required
@permission_required('activos.change_activo', raise_exception=True)
def transferir_activo(request, activo_id=None):
    """Vista para transferir un activo a otro departamento"""
    
    departamentos = [
        "Fiscalización",
        "Recaudación", 
        "Inmuebles Urbanos",
        "Gerencia de Licores",
        "Gerencia General",
        "Jurídica",
        "Informatica",
        "Administración y Finanzas"
        "Calidad de Gestion",
        "GATD"
        
    ]
    
    if request.method == 'GET' and activo_id:
        activo = get_object_or_404(Activo, id=activo_id, activo=True)
        context = {
            'activo': activo,
            'departamentos': [d for d in departamentos if d != activo.departamento],
            'es_admin': request.user.is_superuser,
        }
        return render(request, 'activos/transferir_activo.html', context)
    
    elif request.method == 'POST':
        try:
            data = request.POST
            activo_id = data.get('activo_id')
            activo = get_object_or_404(Activo, id=activo_id, activo=True)
            
            # Guardar departamento anterior antes de cambiar
            departamento_anterior = activo.departamento
            departamento_nuevo = data.get('departamento_destino')
            
            if departamento_anterior == departamento_nuevo:
                messages.warning(request, 'El departamento destino debe ser diferente al actual.')
                return redirect('transferir_activo', activo_id=activo_id)
            
            # Actualizar el activo
            activo.departamento_anterior = departamento_anterior
            activo.departamento = departamento_nuevo
            activo.fecha_transferencia = timezone.now()
            activo.transferido_por = data.get('transferido_por', '').strip()
            activo.recibido_por = data.get('recibido_por', '').strip()
            activo.cargo_entrega = data.get('cargo_entrega', '').strip()
            activo.cargo_recibe = data.get('cargo_recibe', '').strip()
            activo.save()
            
            # Crear registro en historial de transferencias
            transferencia = TransferenciaActivo.objects.create(
                activo=activo,
                departamento_origen=departamento_anterior,
                departamento_destino=departamento_nuevo,
                transferido_por=data.get('transferido_por', '').strip(),
                recibido_por=data.get('recibido_por', '').strip(),
                cargo_entrega=data.get('cargo_entrega', '').strip(),
                cargo_recibe=data.get('cargo_recibe', '').strip(),
                observaciones=data.get('observaciones', '').strip(),
                usuario_registro=request.user.username
            )
            
            messages.success(request, f'✅ Activo "{activo.codigo}" transferido exitosamente.')
            
            # Redirigir a la vista de generación de PDF
            return redirect('generar_acta_transferencia', transferencia_id=transferencia.id)
            
        except Exception as e:
            messages.error(request, f'❌ Error al transferir el activo: {str(e)}')
            return redirect('dashboard')
    
    # Si no es GET con ID ni POST, mostrar lista de activos para transferir
    if request.user.is_superuser:
        activos = Activo.objects.filter(activo=True)
    else:
        activos = Activo.objects.filter(
            activo=True,
            departamento=request.user.departamento
        )
    
    context = {
        'activos': activos,
        'departamentos': departamentos,
        'es_admin': request.user.is_superuser,
    }
    return render(request, 'activos/lista_transferir.html', context)


@login_required
def generar_acta_transferencia(request, transferencia_id):
    """Generar PDF del acta de transferencia"""
    transferencia = get_object_or_404(TransferenciaActivo, id=transferencia_id)
    activo = transferencia.activo
    
    # Crear respuesta HTTP con PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"acta_transferencia_{activo.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Crear el documento PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    estilo_centrado = ParagraphStyle(
        name='Centrado',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=10,
        fontName='Times-Roman'
    )

    # Cargar logos
    logo1_path = os.path.join(settings.BASE_DIR, "activos/static/activos/img/logo1.jpg")
    logo2_path = os.path.join(settings.BASE_DIR, "activos/static/activos/img/slider1.jpg")

    logo1 = Image(logo1_path, width=70, height=70) if os.path.exists(logo1_path) else Paragraph("LOGO 1", estilo_centrado)
    logo2 = Image(logo2_path, width=70, height=70) if os.path.exists(logo2_path) else Paragraph("LOGO 2", estilo_centrado)

    texto_encabezado = Paragraph("""
        <b>República Bolivariana de Venezuela</b><br/>
        <b>Estado Lara</b><br/>
        <b>Alcaldía Bolivariana del Municipio Iribarren</b><br/>
        <b>Servicio Municipal de Administración Tributaria</b><br/>
        <b>SEMAT</b>
    """, estilo_centrado)

    encabezado = Table([[logo1, texto_encabezado, logo2]], colWidths=[80, 400, 80])
    encabezado.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(encabezado)
    elements.append(Spacer(1, 20))

    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=20,
        alignment=1  # Centrado
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=10,
        fontName='Times-Roman',
        textDecoration='underline'
    )
    
    # Estilo para texto normal
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=6,
        fontName='Times-Roman'
    )
    
    # Título del documento
    elements.append(Paragraph("<b>ACTA DE ENTREGA - TRANSFERENCIA DE BIENES</b>", title_style))
    elements.append(Spacer(1, 10))
    

     # OBTENER FECHA CON MES EN ESPAÑOL
    fecha_hora_local = timezone.localtime(timezone.now())
    
    # Diccionario de traducción de meses
    meses_espanol = {
        'January': 'Enero',
        'February': 'Febrero',
        'March': 'Marzo',
        'April': 'Abril',
        'May': 'Mayo',
        'June': 'Junio',
        'July': 'Julio',
        'August': 'Agosto',
        'September': 'Septiembre',
        'October': 'Octubre',
        'November': 'Noviembre',
        'December': 'Diciembre'
    }
    
    # Formatear y traducir
    mes_ingles = fecha_hora_local.strftime('%B')
    dia = fecha_hora_local.strftime('%d')
    anio = fecha_hora_local.strftime('%Y')
    hora = fecha_hora_local.strftime('%I:%M %p')
    mes_espanol = meses_espanol.get(mes_ingles, mes_ingles)
    
    # Construir la cadena final
    hora_fecha = f"{hora} del dia {dia} de {mes_espanol} del {anio}  "

    # Encabezado con fecha y hora
  
    descripcion = f"""
    Siendo las {hora_fecha} se levanta la presente acta para dejar constancia de la transferencia del Bien Público 
    del departamento {transferencia.departamento_origen} al departamento 
    {transferencia.departamento_destino}, con las siguientes características:
    """
    elements.append(Paragraph(descripcion, normal_style))
    elements.append(Spacer(1, 20))
    
    # Tabla con información del activo
    activo_data = [
    ['CÓDIGO', 'SERIAL', 'DESCRIPCIÓN DEL BIEN', 'CONDICIÓN DEL BIEN'],
    [
        activo.codigo,
        activo.serial or 'N/A',
        activo.descripcion or 'Sin descripción adicional',
        activo.condicion.upper()
    ]
]

    # Crear tabla con 4 columnas
    # Ajusta los anchos según lo que necesites
    tabla_activo = Table(activo_data, colWidths=[1*inch, 1.5*inch, 2*inch, 2*inch])
    tabla_activo.setStyle(TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ('ALIGN', (2, 1), (2, 1), 'LEFT'),  # Alinear descripción a la izquierda
    ('FONTNAME', (0, 0), (-1, 0), 'Times-Roman'),
    ('FONTSIZE', (0, 0), (-1, 0), 10),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir ajuste de texto
]))
    
    
    
    elements.append(tabla_activo)
    elements.append(Spacer(1, 20))
    
    # Observaciones si existen
    if transferencia.observaciones:
        elements.append(Paragraph("<b>Observaciones:</b>", normal_style))
        elements.append(Paragraph(transferencia.observaciones, normal_style))
        elements.append(Spacer(1, 20))
    
    # Texto de constancia
    constancia = """
    Dejando constancia de lo antes expuesto y agradeciendo la atención prestada, se procede a la firma del presente documento.
    """
    elements.append(Paragraph(constancia, normal_style))
    elements.append(Spacer(1, 20))
    
    # Tabla de firmas
   # ALTERNATIVA: Tabla de firmas con múltiples filas
    firmas_data = [
    ['ENTREGA CONFORME:', 'RECIBE CONFORME:'],
    [f"Nombre y Apellido:{transferencia.transferido_por}", 
     f"Nombre y Apellido:{transferencia.recibido_por}"],
    [f"Cargo:{transferencia.cargo_entrega}", 
     f"Cargo:{transferencia.cargo_recibe}"],
    [f"Departamento:{transferencia.departamento_origen}", 
     f"Departamento:{transferencia.departamento_destino}"],
    ["Firma:___________________________", 
     "Firma:___________________________"]
]

    tabla_firmas = Table(firmas_data, colWidths=[3*inch, 3*inch])
    tabla_firmas.setStyle(TableStyle([
    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ('FONTNAME', (0, 0), (-1, 0), 'Times-Roman'),
    ('FONTSIZE', (0, 0), (-1, 0), 10),
    ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('SPAN', (0, 0), (0, 0)),  # Si quieres que los títulos ocupen espacio adicional
    ('SPAN', (1, 0), (1, 0)),
]))
    
    elements.append(tabla_firmas)
    elements.append(Spacer(1, 40))
    
    # Nota final
    nota = """
    <b>NOTA:</b> Dicho documento debe ser remitido al área de Bienes Muebles para la notificación 
    correspondiente del movimiento del bien y su registro en el sistema de gestión de activos.
    """
    elements.append(Paragraph(nota, normal_style))
    
    # Construir PDF
    doc.build(elements)
    
    # Marcar como generado
    transferencia.pdf_generado = True
    transferencia.save()
    
    # Obtener el valor del buffer
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response




